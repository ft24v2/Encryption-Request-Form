from dtk import *
from functionsfile import *
from openpyxl import Workbook, load_workbook
from time import gmtime, strftime

### This is the Work Order Submission Form

def outputform(ClientName, SUNet, Phone, Department, Duedateresponse, cryptform, datedueyo, loaneryesno, acyesno, NotesInput):
# Here we open the template we want to put data into

    if cryptform == "Ready Win":
        spreadsheet = "spreadsheetready.xlsx"
    elif cryptform == "New Win":
        spreadsheet = "spreadsheetnew.xlsx"
    elif cryptform == "Upgrade Win":
        spreadsheet = "spreadsheet.xlsx"
    elif cryptform == "Encrypted Win":
        spreadsheet = "spreadsheetcrypt.xlsx"
    else:
        spreadsheet = "spreadsheetready.xlsx"
    

    

    # Specify el spreadsheet to open and copy
    wb2 = load_workbook(spreadsheet)
    ws = wb2.active
    
    # Here we make some case scenarios for departments entered
    if Department == "IT Services":
        Dept = "ITS"
    elif Department == "Earth Sciences":
        Dept = "ES"
    elif Department == "FAO":
        Dept = "UGA"
    elif Department == "Financial Aid":
        Dept = "UGA"
    elif Department == "DAPER":
        Dept = "ATH"
    elif Department == "Daper":
        Dept = "ATH"
    elif Department == "FSI":
        Dept = "FSI"
    elif Department == "UGA":
        Dept = "UGA"
    else:
        Dept = "(____)"
    newnamepre = "{}-{}".format(Dept, getserialnumber) 
    newname = str(newnamepre)
    

    #Date Conversion
    global dateconvertcurrent
    dateconvertcurrent = strftime('%b %d, %Y')
    
# Here is where the data gets entered in the cell
    ws['C2'] = ClientName
    ws['E2'] = SUNet
    ws['C3'] = Phone
    ws['E3'] = Department
    ws['C5'] = str(hostname)
    ws['E5'] = newname
    ws['C6'] = dateconvertcurrent
    ws['E6'] = datedueyo
    ws['E4'] = loaneryesno
    ws['C4'] = acyesno
    ws['C7'] = NotesInput
   

# Here we save the data into a new work book, and specify where it goes. YEAH BOI!!!!!!

    if not os.path.exists('..\\WorkOrders\\{}-{}-{}'.format(ClientName, Department, str(hostname))):
        os.makedirs('..\\WorkOrders\\{}-{}-{}'.format(ClientName, Department, str(hostname)))
    wb2.save('..\\WorkOrders\\{}-{}-{}\\{}-{}-WorkOrder.xlsx'.format(ClientName, Department, str(hostname), ClientName, str(hostname)))


### This is the System Info Grabber Form, The "Print" Function will eventually become WRITE, and will write
### Out a log file in the same dir as the worksheets.
